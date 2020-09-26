import pymysql
from openpyxl import workbook
from openpyxl import load_workbook

class MysqlController:
    def __init__(self, host, id, pw, db_name):
        try:
            self.conn = pymysql.connect(host=host, user=id,password=pw,db=db_name) # db와 연결.
            self.curs = self.conn.cursor() # 접속 성공시, cursor객체를 가져옴.
        except:
            print('connection error')

    def select_all(self):
        try:
            with self.curs as cur:
                sql = "select * from Apple" 
                cur.execute(sql)
                rs = cur.fetchall() # 쿼리 결과의 모든 행을 검색하여 순서대로 반환한다.
                for row in rs:
                    print(row)
        except:
            print('error')
 
    def insert_excel_to_db(self):
        try:
            sql = 'insert into Apple value(%s,%s,%s,%s,%s,%s,%s)'
            print('sql')
            wb = load_workbook(r'C:/Users/chank/Python38-32/크롤링 자료/study/Crawling_Study/week5/Apple_stock.xlsx',data_only=True)
            # excel파일 open 후 wb에 저장.
            ws = wb['Sheet1'] 
            # Sheet1의 data를 가져옴.
            print('success open excel')
            
            iter_rows = iter(ws.rows) # ws.rows 객체의 반복자를 가져옴.  
            next(iter_rows) # 첫번째 행 생략
            for row in iter_rows:
                self.curs.execute(sql,(row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value))
            
            print('insert content')
            self.conn.commit() # 수행 결과 commit
        except:
            print('insert error')

if __name__ == "__main__":
    db = MysqlController('localhost','root','skdltm97','test')
    # db.insert_excel_to_db()
    db.select_all()

